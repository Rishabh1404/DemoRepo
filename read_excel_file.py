import openpyxl

dt = openpyxl.load_workbook(r"C:\Users\Rishabh_Kumar\Documents\StudentData.xlsx")
sheet_obj = dt.active
max_column = sheet_obj.max_column
max_row = sheet_obj.max_row

print("Values of excel sheet are below")
for i in range(1,max_column+1):
    for j in range(1,max_row+1):
        cell_value = sheet_obj.cell(row=j, column=i).value
        print(cell_value, end=' ')

